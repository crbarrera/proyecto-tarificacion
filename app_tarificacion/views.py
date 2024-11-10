from datetime import timedelta
from datetime import datetime
from django.forms import ValidationError
from django.shortcuts import redirect, render, get_object_or_404
from weasyprint import HTML
from .models import Anexo, Codunidad, Ctapresu, Llamada, Tarificacion, Usuario, Proveedor
from .forms import AnexoForm, CuentaPresupuestariaForm, LoginForm, UsuarioForm, ProveedorForm, CodigoUnidadForm
from django.db import IntegrityError, connection, DatabaseError
from django.contrib.auth import authenticate, login as auth_login
from django.shortcuts import redirect
from django.contrib.auth import login as auth_login, authenticate  
from django.shortcuts import render
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum, F, Avg, ExpressionWrapper, fields
import cx_Oracle
from django.conf import settings
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.views import LogoutView
from django.contrib.auth.hashers import make_password
from django.utils.timezone import make_aware, get_current_timezone
import pytz
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
import base64
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment
import os
from decimal import Decimal
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from io import BytesIO
from openpyxl.cell.cell import MergedCell


@login_required
def index(request):
    rol_user = None
    if request.user.is_authenticated:
        rol_user = request.user.rol_usuario  # Accede al rol del usuario autenticado
    
    return render(request, 'index.html', {'rol_user': rol_user})
  

# def login_view(request):
#     if request.method == 'POST':
#         username = request.POST.get('username')
#         password = request.POST.get('password')

#         # Autenticación del usuario
#         user = authenticate(request, username=username, password=password)

#         if user is not None:
#             auth_login(request, user)  # Usamos auth_login en vez de login para evitar conflictos
#             return redirect(request.GET.get('next', 'home'))  # Redirigir a la página principal o a 'next'
#         else:
#             messages.error(request, 'Nombre de usuario o contraseña incorrectos.')

#     return render(request, 'iniciar_sesion.html')
  
class CustomLogoutView(LogoutView):
    next_page = 'login'  # Asegúrate de que 'login' esté definido en tus urls

@login_required
def listar_usuarios(request):
    query = request.GET.get('query', '')
    django_cursor = connection.cursor()
    cursor = django_cursor.connection.cursor()
    out_cur = django_cursor.connection.cursor()

    if query:
        cursor.callproc('buscar_usuarios', [query, out_cur])
    else:
        cursor.callproc('listar_usuarios', [out_cur])

    usuarios = []

    for fila in out_cur:
        id_usuario, username, email_usuario, rol_usuario = fila
        usuarios.append({
            'id_usuario': id_usuario,
            'username': username,
            'email_usuario': email_usuario,
            'rol_usuario': rol_usuario
        })

    return render(request, 'listar_usuarios.html', {'usuarios': usuarios})

@login_required
def crear_usuario(request):
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            rol_usuario = form.cleaned_data['rol_usuario']
            email_usuario = form.cleaned_data['email_usuario']

            try:
                # Genera el hash de la contraseña con Django
                password_hashed = make_password(password)

                # Llama al procedimiento almacenado pasándole el hash de la contraseña
                with connection.cursor() as cursor:
                    cursor.callproc('crear_usuario', [username, password_hashed, rol_usuario, email_usuario])

                return redirect('usuarios')
            except DatabaseError as e:
                error_message = str(e)
                if 'ORA-20002' in error_message:
                    form.add_error('username', 'El nombre de usuario ya existe.')
                else:
                    form.add_error(None, 'Ocurrió un error al crear el usuario.')
    else:
        form = UsuarioForm()
    
    return render(request, 'crear_usuario.html', {'form': form})

@login_required
def modificar_usuario(request, id):
    # Obtiene el usuario actual basado en el ID proporcionado
    usuario_actual = get_object_or_404(Usuario, id_usuario=id)

    if request.method == 'POST':
        # Se pasa la instancia del usuario actual al formulario
        form = UsuarioForm(request.POST, instance=usuario_actual)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']  # Esto será vacío si no se cambia
            rol_usuario = form.cleaned_data['rol_usuario']
            email_usuario = form.cleaned_data['email_usuario']

            try:
                # Genera un nuevo hash solo si se proporciona una nueva contraseña
                if password:
                    password_hashed = make_password(password)
                else:
                    password_hashed = usuario_actual.password  # Mantiene el hash actual si no se cambia

                # Llama al procedimiento almacenado
                with connection.cursor() as cursor:
                    cursor.callproc('modificar_usuario', [
                        usuario_actual.id_usuario,
                        username,
                        password_hashed,  # Asegúrate de pasar el hash
                        rol_usuario,
                        email_usuario
                    ])
                
                return redirect('usuarios')  # Redirige a la lista de usuarios
            except DatabaseError as e:
                error_message = str(e)
                if 'ORA-20001' in error_message:
                    form.add_error('rol_usuario', 'Rol de usuario inválido.')
                else:
                    form.add_error(None, 'Ocurrió un error al modificar el usuario.')
    else:
        # Prepara el formulario con los datos actuales del usuario
        form = UsuarioForm(instance=usuario_actual)

    return render(request, 'modificar_usuario.html', {'form': form, 'usuario': usuario_actual})

@login_required
def eliminar_usuario(request, id):
    if request.method == 'POST':
        try:
            with connection.cursor() as cursor:
                cursor.callproc('eliminar_usuario', [id])
            return redirect('usuarios')  # Redirige a la lista de usuarios después de eliminar
        except DatabaseError as e:
            # Manejo de errores si es necesario
            error_message = str(e)
            # Puedes manejar el error según lo necesites
            # Por ejemplo, agregar un mensaje de error a la sesión

    return redirect('usuarios')  # Redirige si la solicitud no es POST

@login_required
def listar_proveedores(request):
    query = request.GET.get('query', '')
    django_cursor = connection.cursor()
    cursor = django_cursor.connection.cursor()
    out_cur = django_cursor.connection.cursor()

    if query:
        cursor.callproc('buscar_proveedores', [query, out_cur])
    else:
        cursor.callproc('listar_proveedores', [out_cur])

    proveedores = []

    for fila in out_cur:
        id_proveedor, nombre_proveedor, tarifa_cel, tarifa_slm, tarifa_ldi = fila
        proveedores.append({
            'id_proveedor': id_proveedor,
            'nombre_proveedor': nombre_proveedor,
            'tarifa_cel': tarifa_cel,
            'tarifa_slm': tarifa_slm,
            'tarifa_ldi': tarifa_ldi
        })

    return render(request, 'listar_proveedores.html', {'proveedores': proveedores})
  
@login_required
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            nombre_proveedor = form.cleaned_data['nombre_proveedor']
            tarifa_cel = form.cleaned_data['tarifa_cel']
            tarifa_slm = form.cleaned_data['tarifa_slm']
            tarifa_ldi = form.cleaned_data['tarifa_ldi']
            
            try:
                with connection.cursor() as cursor:
                    cursor.callproc('crear_proveedor', [nombre_proveedor, tarifa_cel, tarifa_slm, tarifa_ldi])
                return redirect('proveedores')
            except DatabaseError as e:
                form.add_error(None, 'Ocurrió un error al crear el proveedor.')
    else:
        form = ProveedorForm()
    
    return render(request, 'crear_proveedor.html', {'form': form})

@login_required
def modificar_proveedor(request, id):
    proveedor_actual = get_object_or_404(Proveedor, id_proveedor=id)

    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            nombre_proveedor = form.cleaned_data['nombre_proveedor']
            tarifa_cel = form.cleaned_data['tarifa_cel']
            tarifa_slm = form.cleaned_data['tarifa_slm']
            tarifa_ldi = form.cleaned_data['tarifa_ldi']

            try:
                with connection.cursor() as cursor:
                    cursor.callproc('modificar_proveedor', [proveedor_actual.id_proveedor, nombre_proveedor, tarifa_cel, tarifa_slm, tarifa_ldi])
                return redirect('proveedores')
            except DatabaseError as e:
                form.add_error(None, 'Ocurrió un error al modificar el proveedor.')
    else:
        initial_data = {
            'nombre_proveedor': proveedor_actual.nombre_proveedor,
            'tarifa_cel': proveedor_actual.tarifa_cel,
            'tarifa_slm': proveedor_actual.tarifa_slm,
            'tarifa_ldi': proveedor_actual.tarifa_ldi,
        }
        form = ProveedorForm(initial=initial_data)

    return render(request, 'modificar_proveedor.html', {'form': form, 'proveedor': proveedor_actual})

@login_required
def eliminar_proveedor(request, id):
    if request.method == 'POST':
        try:
            with connection.cursor() as cursor:
                cursor.callproc('eliminar_proveedor', [id])
            return redirect('proveedores')
        except DatabaseError as e:
            # Manejo de errores si es necesario
            error_message = str(e)
            # Puedes manejar el error según lo necesites
            # Por ejemplo, agregar un mensaje de error a la sesión

    return redirect('proveedores')

@login_required
def listar_codigos_unidad(request):
    query = request.GET.get('query', '')
    django_cursor = connection.cursor()
    cursor = django_cursor.connection.cursor()
    out_cur = django_cursor.connection.cursor()

    if query:
        cursor.callproc('buscar_codigos_unidad', [query, out_cur])
    else:
        cursor.callproc('listar_codigos_unidad', [out_cur])

    codigos_unidad = []

    for fila in out_cur:
        id_codigo, nombre_codigo, ctapresu_id = fila
        ctapresu = Ctapresu.objects.get(id_cuenta=ctapresu_id)
        codigos_unidad.append({
            'id_codigo': id_codigo,
            'nombre_codigo': nombre_codigo,
            'ctapresu': ctapresu.nombre_cuenta
        })

    return render(request, 'listar_codigos_unidad.html', {'codigos_unidad': codigos_unidad})

@login_required
def crear_codigo_unidad(request):
    if request.method == 'POST':
        form = CodigoUnidadForm(request.POST)
        if form.is_valid():
            nombre_codigo = form.cleaned_data['nombre_codigo']
            ctapresu_id = form.cleaned_data['ctapresu_id']
            
            try:
                with connection.cursor() as cursor:
                    cursor.callproc('crear_codigo_unidad', [nombre_codigo, ctapresu_id.id_cuenta])
                return redirect('codigos_unidad')
            except DatabaseError as e:
                form.add_error(None, 'Ocurrió un error al crear el código de unidad.')
        else:
            print(form.errors)  # Agrega esto para depurar errores de validación
    else:
        form = CodigoUnidadForm()
    
    return render(request, 'crear_codigo_unidad.html', {'form': form})

@login_required
def modificar_codigo_unidad(request, id):
    codigo_unidad_actual = get_object_or_404(Codunidad, id_codigo=id)

    if request.method == 'POST':
        form = CodigoUnidadForm(request.POST)
        if form.is_valid():
            nombre_codigo = form.cleaned_data['nombre_codigo']
            ctapresu_id = form.cleaned_data['ctapresu_id']

            try:
                with connection.cursor() as cursor:
                    cursor.callproc('modificar_codigo_unidad', [codigo_unidad_actual.id_codigo, nombre_codigo, ctapresu_id.id_cuenta])
                return redirect('codigos_unidad')
            except DatabaseError as e:
                form.add_error(None, 'Ocurrió un error al modificar el código de unidad.')
    else:
        initial_data = {
            'nombre_codigo': codigo_unidad_actual.nombre_codigo,
            'ctapresu_id': codigo_unidad_actual.ctapresu
        }
        form = CodigoUnidadForm(initial=initial_data)

    return render(request, 'modificar_codigo_unidad.html', {'form': form, 'codigo_unidad': codigo_unidad_actual})

@login_required
def eliminar_codigo_unidad(request, id):
    if request.method == 'POST':
        try:
            with connection.cursor() as cursor:
                cursor.callproc('eliminar_codigo_unidad', [id])
            return redirect('codigos_unidad')
        except DatabaseError as e:
            # Manejo de errores si es necesario
            error_message = str(e)
            # Puedes manejar el error según lo necesites
            # Por ejemplo, agregar un mensaje de error a la sesión

    return redirect('codigos_unidad')

@login_required
def listar_cuentas_presupuestarias(request):
    query = request.GET.get('query', '')
    django_cursor = connection.cursor()
    cursor = django_cursor.connection.cursor()
    out_cur = django_cursor.connection.cursor()

    if query:
        cursor.callproc('buscar_cuentas_presupuestarias', [query, out_cur])
    else:
        cursor.callproc('listar_cuentas_presupuestarias', [out_cur])

    cuentas_presupuestarias = []

    for fila in out_cur:
        id_cuenta, nombre_cuenta = fila
        cuentas_presupuestarias.append({
            'id_cuenta': id_cuenta,
            'nombre_cuenta': nombre_cuenta
        })

    return render(request, 'listar_cuentas_presupuestarias.html', {'cuentas_presupuestarias': cuentas_presupuestarias})

@login_required
def crear_cuenta_presupuestaria(request):
    if request.method == 'POST':
        form = CuentaPresupuestariaForm(request.POST)
        if form.is_valid():
            nombre_cuenta = form.cleaned_data['nombre_cuenta']
            
            try:
                with connection.cursor() as cursor:
                    cursor.callproc('crear_cuenta_presupuestaria', [nombre_cuenta])
                return redirect('cuentas_presupuestarias')
            except DatabaseError as e:
                form.add_error(None, 'Ocurrió un error al crear la cuenta presupuestaria.')
    else:
        form = CuentaPresupuestariaForm()
    
    return render(request, 'crear_cuenta_presupuestaria.html', {'form': form})

@login_required
def modificar_cuenta_presupuestaria(request, id):
    cuenta_actual = get_object_or_404(Ctapresu, id_cuenta=id)

    if request.method == 'POST':
        form = CuentaPresupuestariaForm(request.POST)
        if form.is_valid():
            nombre_cuenta = form.cleaned_data['nombre_cuenta']

            try:
                with connection.cursor() as cursor:
                    cursor.callproc('modificar_cuenta_presupuestaria', [cuenta_actual.id_cuenta, nombre_cuenta])
                return redirect('cuentas_presupuestarias')
            except DatabaseError as e:
                form.add_error(None, 'Ocurrió un error al modificar la cuenta presupuestaria.')
    else:
        initial_data = {
            'nombre_cuenta': cuenta_actual.nombre_cuenta,
        }
        form = CuentaPresupuestariaForm(initial=initial_data)

    return render(request, 'modificar_cuenta_presupuestaria.html', {'form': form, 'cuenta': cuenta_actual})

@login_required
def eliminar_cuenta_presupuestaria(request, id):
    if request.method == 'POST':
        try:
            with connection.cursor() as cursor:
                cursor.callproc('eliminar_cuenta_presupuestaria', [id])
            return redirect('cuentas_presupuestarias')
        except DatabaseError as e:
            # Manejo de errores si es necesario
            error_message = str(e)
            # Puedes manejar el error según lo necesites
            # Por ejemplo, agregar un mensaje de error a la sesión

    return redirect('cuentas_presupuestarias')

@login_required
def listar_anexos(request):
    query = request.GET.get('query', '')
    django_cursor = connection.cursor()
    cursor = django_cursor.connection.cursor()
    out_cur = django_cursor.connection.cursor()

    if query:
        cursor.callproc('buscar_anexos', [query, out_cur])
    else:
        cursor.callproc('listar_anexos', [out_cur])

    anexos = []

    for fila in out_cur:
        id_anexo, numero_anexo, cargo_fijo, estado_anexo, usuario_id, codunidad_id = fila
        usuario = Usuario.objects.get(id_usuario=usuario_id)
        codunidad = Codunidad.objects.get(id_codigo=codunidad_id)
        anexos.append({
            'id_anexo': id_anexo,
            'numero_anexo': numero_anexo,
            'cargo_fijo': cargo_fijo,
            'estado_anexo': estado_anexo,
            'usuario': usuario.username,
            'codunidad': codunidad.nombre_codigo
        })

    return render(request, 'listar_anexos.html', {'anexos': anexos})

@login_required
def crear_anexo(request):
    if request.method == 'POST':
        form = AnexoForm(request.POST)
        if form.is_valid():
            numero_anexo = form.cleaned_data['numero_anexo']
            cargo_fijo = form.cleaned_data['cargo_fijo']
            estado_anexo = form.cleaned_data['estado_anexo']
            usuario = form.cleaned_data['usuario_id']
            codunidad = form.cleaned_data['codunidad_id']
            
            try:
                with connection.cursor() as cursor:
                    cursor.callproc('crear_anexo', [numero_anexo, cargo_fijo, estado_anexo, usuario.id_usuario, codunidad.id_codigo])
                return redirect('anexos')
            except DatabaseError as e:
                form.add_error(None, 'Ocurrió un error al crear el anexo.')
        else:
            print(form.errors)  # Agrega esto para depurar errores de validación
    else:
        form = AnexoForm()
    
    return render(request, 'crear_anexo.html', {'form': form})

@login_required
def modificar_anexo(request, id):
    anexo_actual = get_object_or_404(Anexo, id_anexo=id)

    if request.method == 'POST':
        form = AnexoForm(request.POST)
        if form.is_valid():
            numero_anexo = form.cleaned_data['numero_anexo']
            cargo_fijo = form.cleaned_data['cargo_fijo']
            estado_anexo = form.cleaned_data['estado_anexo']
            usuario = form.cleaned_data['usuario_id']
            codunidad = form.cleaned_data['codunidad_id']

            try:
                with connection.cursor() as cursor:
                    cursor.callproc('modificar_anexo', [anexo_actual.id_anexo, numero_anexo, cargo_fijo, estado_anexo, usuario.id_usuario, codunidad.id_codigo])
                return redirect('anexos')
            except DatabaseError as e:
                form.add_error(None, 'Ocurrió un error al modificar el anexo.')
    else:
        initial_data = {
            'numero_anexo': anexo_actual.numero_anexo,
            'cargo_fijo': anexo_actual.cargo_fijo,
            'estado_anexo': anexo_actual.estado_anexo,
            'usuario_id': anexo_actual.usuario,
            'codunidad_id': anexo_actual.codunidad,
        }
        form = AnexoForm(initial=initial_data)

    return render(request, 'modificar_anexo.html', {'form': form, 'anexo': anexo_actual})

@login_required
def eliminar_anexo(request, id):
    if request.method == 'POST':
        try:
            with connection.cursor() as cursor:
                cursor.callproc('eliminar_anexo', [id])
            return redirect('anexos')
        except DatabaseError as e:
            # Manejo de errores si es necesario
            error_message = str(e)
            # Puedes manejar el error según lo necesites
            # Por ejemplo, agregar un mensaje de error a la sesión

    return redirect('anexos')

@login_required
def listar_responsables_unidad(request):
    query = request.GET.get('query', '')
    django_cursor = connection.cursor()
    cursor = django_cursor.connection.cursor()
    out_cur = django_cursor.connection.cursor()

    if query:
        cursor.callproc('buscar_responsables_unidad', [query, out_cur])
    else:
        cursor.callproc('listar_responsables_unidad', [out_cur])

    responsables = []

    for fila in out_cur:
        id_usuario, username, email_usuario, rol_usuario = fila
        responsables.append({
            'id_usuario': id_usuario,
            'username': username,
            'email_usuario': email_usuario,
            'rol_usuario': rol_usuario
        })

    return render(request, 'listar_responsables_unidad.html', {'responsables': responsables})

@login_required
def listar_anexos_activos(request):
    query = request.GET.get('query', '')
    django_cursor = connection.cursor()
    cursor = django_cursor.connection.cursor()
    out_cur = django_cursor.connection.cursor()

    if query:
        cursor.callproc('buscar_anexos_activos', [query, out_cur])
    else:
        cursor.callproc('listar_anexos_activos', [out_cur])

    anexos = []

    for fila in out_cur:
        id_anexo, numero_anexo, cargo_fijo, estado_anexo, usuario_id, codunidad_id = fila
        usuario = Usuario.objects.get(id_usuario=usuario_id)
        codunidad = Codunidad.objects.get(id_codigo=codunidad_id)
        anexos.append({
            'id_anexo': id_anexo,
            'numero_anexo': numero_anexo,
            'cargo_fijo': cargo_fijo,
            'estado_anexo': estado_anexo,
            'usuario': usuario.username,
            'codunidad': codunidad.nombre_codigo
        })

    return render(request, 'listar_anexos_activos.html', {'anexos': anexos})

@login_required
def calcular_tarificacion(request):
    mensaje = None
    if request.method == 'POST':
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_termino = request.POST.get('fecha_termino')
        anexo_id = request.POST.get('anexo')

        # Convertir las fechas a formato date (sin hora)
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_termino = datetime.strptime(fecha_termino, "%Y-%m-%d").date()
        except ValueError:
            mensaje = 'Formato de fecha incorrecto'
            anexos = Anexo.objects.all()
            return render(request, 'calcular_tarificacion.html', {'anexos': anexos, 'mensaje': mensaje})

        # Obtener el anexo seleccionado
        anexo = get_object_or_404(Anexo, id_anexo=anexo_id)

        # Obtener todas las llamadas de este anexo en el rango de fechas
        llamadas = Llamada.objects.filter(
            anexo=anexo,
            inicio_llamada__range=[fecha_inicio, fecha_termino]
        ).select_related('proveedor')

        # Inicializar variables para sumar el costo y los minutos
        costo_total = 0
        minutos_totales = 0
        tipo_llamada = None  # Inicializar tipo_llamada

        # Procesar cada llamada
        for llamada in llamadas:
            # Obtener tarifa según el tipo de llamada
            if llamada.tipo_llamada == 'CEL':
                tarifa = llamada.proveedor.tarifa_cel
            elif llamada.tipo_llamada == 'SLM':
                tarifa = llamada.proveedor.tarifa_slm
            elif llamada.tipo_llamada == 'LDI':
                tarifa = llamada.proveedor.tarifa_ldi
            else:
                tarifa = 0

            # Calcular el costo de la llamada y sumar al total
            costo_llamada = llamada.segundos_facturados * tarifa
            costo_total += costo_llamada

            # Sumar minutos (convertidos de segundos)
            minutos_totales += llamada.segundos_facturados / 60

            # Asignar el tipo de llamada, puede ser el último o puedes usar otro criterio
            tipo_llamada = llamada.tipo_llamada  # Asignamos el tipo de llamada de la última llamada

        # Sumar el cargo fijo al costo total
        costo_total += anexo.cargo_fijo

        # Solo crear el registro si hay llamadas o si el anexo tiene cargo fijo
        Tarificacion.objects.create(
            anexo=anexo,
            fecha_inicio=fecha_inicio,
            fecha_termino=fecha_termino,
            costo_total=costo_total,
            minutos_totales=int(minutos_totales),  # Convertir a entero
            tipo_llamada=tipo_llamada if tipo_llamada else 'N/A'  # Asignar un valor por defecto si no hay llamadas
        )

        mensaje = 'Tarificación calculada correctamente'

    # Renderizar el formulario si el método es GET o después de procesar el POST
    anexos = Anexo.objects.all()
    return render(request, 'calcular_tarificacion.html', {'anexos': anexos, 'mensaje': mensaje})

@login_required
def listar_tarificacion(request):
    # Crea un cursor de conexión a la base de datos
    django_cursor = connection.cursor()
    out_cur = django_cursor.connection.cursor()  # Cursor de salida para el procedimiento almacenado

    # Inicializa una lista para almacenar los resultados
    tarificaciones = []

    try:
        # Llama al procedimiento almacenado
        django_cursor.callproc('listar_tarificaciones', [out_cur])  # Asegúrate de que el nombre del procedimiento esté correcto

        # Itera sobre el cursor de salida
        for fila in out_cur:
            # Desempaqueta los valores del cursor de salida
            id_tarificacion, fecha_inicio, fecha_termino, costo_total, minutos_totales, tipo_llamada, anexo_id = fila

            # Agrega cada tarificación a la lista
            tarificaciones.append({
                'id_tarificacion': id_tarificacion,
                'fecha_inicio': fecha_inicio,
                'fecha_termino': fecha_termino,
                'costo_total': costo_total,
                'minutos_totales': minutos_totales,
                'tipo_llamada': tipo_llamada,
                'anexo_id': anexo_id,  # Guarda el ID del anexo
            })
    except Exception as e:
        print("Error al obtener datos: ", e)  # Imprime cualquier error

    # Obtén todos los anexos para la lista desplegable
    anexos = Anexo.objects.all()

    # Filtra las tarificaciones si se selecciona un anexo
    anexo_id = request.GET.get('anexo_id')
    if anexo_id:
        tarificaciones = [t for t in tarificaciones if t['anexo_id'] == int(anexo_id)]

    # Renderiza la plantilla y pasa los datos de tarificaciones y anexos
    return render(request, 'listar_tarificacion.html', {'tarificaciones': tarificaciones, 'anexos': anexos})

@login_required
def listar_tarificacion_por_codunidad(request):
    django_cursor = connection.cursor()
    out_cur = django_cursor.connection.cursor()

    tarificaciones = []

    try:
        django_cursor.callproc('listar_tarificaciones', [out_cur])

        for fila in out_cur:
            id_tarificacion, fecha_inicio, fecha_termino, costo_total, minutos_totales, tipo_llamada, anexo_id = fila
            tarificaciones.append({
                'id_tarificacion': id_tarificacion,
                'fecha_inicio': fecha_inicio,
                'fecha_termino': fecha_termino,
                'costo_total': costo_total,
                'minutos_totales': minutos_totales,
                'tipo_llamada': tipo_llamada,
                'anexo_id': anexo_id,
            })
    except Exception as e:
        print("Error al obtener datos: ", e)

    codunidades = Codunidad.objects.all()

    codunidad_id = request.GET.get('codunidad_id')

    if codunidad_id:
        try:
            codunidad_id = int(codunidad_id)
            tarificaciones = [t for t in tarificaciones if Anexo.objects.get(id_anexo=t['anexo_id']).codunidad_id == codunidad_id]
        except ValueError:
            pass

    return render(request, 'listar_tarificacion_por_codigo.html', {'tarificaciones': tarificaciones, 'codunidades': codunidades})

@login_required
def consultar_tarificacion_anexo(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_termino = request.GET.get('fecha_termino')
    anexo_id = request.GET.get('anexo')

    anexos = Anexo.objects.all()  # Obtener todos los anexos
    rol_user = None

    if request.user.is_authenticated:
        rol_user = request.user.rol_usuario  # Accede al rol del usuario autenticado

    if fecha_inicio and fecha_termino and anexo_id:
        try:
            # Convertir las fechas a objetos datetime
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_termino_dt = datetime.strptime(fecha_termino, '%Y-%m-%d')

            if fecha_inicio_dt > fecha_termino_dt:
                raise ValidationError("La fecha de inicio debe ser anterior a la fecha de término.")

            # Imprimir valores para depuración
            print(f"Fecha Inicio: {fecha_inicio_dt}")
            print(f"Fecha Término: {fecha_termino_dt}")
            print(f"Anexo ID: {anexo_id}")

            tarificaciones = Tarificacion.objects.filter(
                fecha_inicio__gte=fecha_inicio_dt,
                fecha_termino__lte=fecha_termino_dt,
                anexo_id=anexo_id
            ).select_related('anexo')

            # Imprimir la cantidad de registros encontrados
            print(f"Tarificaciones encontradas: {tarificaciones.count()}")

            context = {
                'tarificaciones': tarificaciones,
                'anexos': anexos,  # Pasar los anexos al contexto
                'rol_user': rol_user  # Pasar el rol del usuario al contexto
            }
            return render(request, 'listar_tarificacion_anexo.html', context)
        except (ValueError, ValidationError) as e:
            context = {
                'anexos': anexos,
                'error': str(e),
                'rol_user': rol_user  # Pasar el rol del usuario al contexto
            }
            return render(request, 'consultar_tarificacion_anexo.html', context)
    else:
        context = {
            'anexos': anexos,  # Pasar los anexos al contexto
            'rol_user': rol_user  # Pasar el rol del usuario al contexto
        }
        return render(request, 'consultar_tarificacion_anexo.html', context)

@login_required   
def consultar_tarificacion_codigo(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_termino = request.GET.get('fecha_termino')
    codigo_id = request.GET.get('codigo')

    codigos = Codunidad.objects.all()  # Obtener todos los códigos de unidad
    rol_user = None

    if request.user.is_authenticated:
        rol_user = request.user.rol_usuario  # Accede al rol del usuario autenticado

    if fecha_inicio and fecha_termino and codigo_id:
        try:
            # Convertir las fechas a objetos datetime
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_termino_dt = datetime.strptime(fecha_termino, '%Y-%m-%d')

            if fecha_inicio_dt > fecha_termino_dt:
                raise ValidationError("La fecha de inicio debe ser anterior a la fecha de término.")

            # Imprimir valores para depuración
            print(f"Fecha Inicio: {fecha_inicio_dt}")
            print(f"Fecha Término: {fecha_termino_dt}")
            print(f"Código ID: {codigo_id}")

            # Ajustar la consulta para incluir el día completo de la fecha de término
            tarificaciones = Tarificacion.objects.filter(
                fecha_inicio__gte=fecha_inicio_dt,
                fecha_termino__lte=fecha_termino_dt + timedelta(days=1),
                anexo__codunidad_id=codigo_id
            ).select_related('anexo', 'anexo__codunidad')

            # Imprimir la cantidad de registros encontrados
            print(f"Tarificaciones encontradas: {tarificaciones.count()}")

            context = {
                'tarificaciones': tarificaciones,
                'codigos': codigos,  # Pasar los códigos al contexto
                'rol_user': rol_user  # Pasar el rol del usuario al contexto
            }
            return render(request, 'listar_tarificacion_codigo.html', context)
        except (ValueError, ValidationError) as e:
            context = {
                'codigos': codigos,
                'error': str(e),
                'rol_user': rol_user  # Pasar el rol del usuario al contexto
            }
            return render(request, 'consultar_tarificacion_codigo.html', context)
    else:
        context = {
            'codigos': codigos,  # Pasar los códigos al contexto
            'rol_user': rol_user  # Pasar el rol del usuario al contexto
        }
        return render(request, 'consultar_tarificacion_codigo.html', context)

@login_required
def detalle_tarificacion(request, id_tarificacion):
    tarificacion = get_object_or_404(Tarificacion, id_tarificacion=id_tarificacion)
    referer = request.META.get('HTTP_REFERER', '/')
    rol_user = None

    if request.user.is_authenticated:
        rol_user = request.user.rol_usuario  # Accede al rol del usuario autenticado

    return render(request, 'detalle_tarificacion.html', {
        'tarificacion': tarificacion,
        'referer': referer,
        'rol_user': rol_user  # Pasar el rol del usuario al contexto
    })

@login_required
def generar_reporte_pdf(request, id_tarificacion):
    tarificacion = get_object_or_404(Tarificacion, id_tarificacion=id_tarificacion)
    
    # Renderizar la plantilla HTML con los datos de la tarificación
    html_string = render_to_string('reporte_tarificacion.html', {'tarificacion': tarificacion})
    
    # Generar el PDF
    html = HTML(string=html_string)
    pdf = html.write_pdf()

    # Crear la respuesta HTTP con el PDF
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_tarificacion_{id_tarificacion}.pdf"'
    
    return response
  
def generar_reporte_xls(request, id_tarificacion):
    tarificacion = get_object_or_404(Tarificacion, id_tarificacion=id_tarificacion)
    
    # Crear un libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Tarificación"
    
    # Insertar una fila vacía al inicio
    ws.append(['', ''])
    
    # Añadir el título y fusionar celdas
    title = "Reporte de Tarificación"
    ws.merge_cells('B2:C2')
    cell = ws['B2']
    cell.value = title
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Añadir los encabezados con una columna vacía al inicio
    headers = ['', 'Campo', 'Valor']
    ws.append(headers)
    
    # Añadir los datos de la tarificación con una columna vacía al inicio
    data = [
        ['', 'Id. Tarificación', tarificacion.id_tarificacion],
        ['', 'Fecha Inicio', tarificacion.fecha_inicio.strftime('%d/%m/%Y')],
        ['', 'Fecha Término', tarificacion.fecha_termino.strftime('%d/%m/%Y')],
        ['', 'Costo Total', tarificacion.costo_total],
        ['', 'Minutos Totales', tarificacion.minutos_totales],
        ['', 'Código de Unidad', tarificacion.anexo.codunidad.nombre_codigo]
    ]
    
    for row in data:
        ws.append(row)
    
    # Ajustar el ancho de las columnas y alinear el contenido
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Obtener la letra de la columna
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            # Alinear el contenido al centro
            cell.alignment = Alignment(horizontal='center')
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Añadir bordes a las celdas
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.border = thin_border
    
    # Crear la respuesta HTTP con el archivo XLS
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_tarificacion_{id_tarificacion}.xlsx"'
    
    # Guardar el libro de trabajo en la respuesta
    wb.save(response)
    
    return response

@login_required
def enviar_reporte(request, id_tarificacion):
    tarificacion = get_object_or_404(Tarificacion, id_tarificacion=id_tarificacion)
    
    # Generar el PDF
    html_string = render_to_string('reporte_tarificacion.html', {'tarificacion': tarificacion})
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    
    # Generar el XLS
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Tarificación"
    
    ws.append(['', ''])
    title = "Reporte de Tarificación"
    ws.merge_cells('B2:C2')
    cell = ws['B2']
    cell.value = title
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    headers = ['', 'Campo', 'Valor']
    ws.append(headers)
    
    data = [
        ['', 'Id. Tarificación', tarificacion.id_tarificacion],
        ['', 'Fecha Inicio', tarificacion.fecha_inicio.strftime('%d/%m/%Y')],
        ['', 'Fecha Término', tarificacion.fecha_termino.strftime('%d/%m/%Y')],
        ['', 'Costo Total', tarificacion.costo_total],
        ['', 'Minutos Totales', tarificacion.minutos_totales],
        ['', 'Código de Unidad', tarificacion.anexo.codunidad.nombre_codigo]
    ]
    
    for row in data:
        ws.append(row)
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            cell.alignment = Alignment(horizontal='center')
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.border = thin_border
    
    xls_io = BytesIO()
    wb.save(xls_io)
    xls_io.seek(0)
    xls_data = xls_io.read()
    
    # Codificar los archivos en base64
    encoded_pdf = base64.b64encode(pdf).decode()
    encoded_xls = base64.b64encode(xls_data).decode()
    
    # Crear el mensaje de correo
    message = Mail(
        from_email='cr.barrera@duocuc.cl',
        to_emails='cr.barrera@duocuc.cl',
        subject='Reporte de Tarificación',
        html_content='Adjunto encontrarás el reporte de tarificación solicitado.'
    )
    
    # Adjuntar el PDF
    pdf_attachment = Attachment(
        file_content=encoded_pdf,
        file_type='application/pdf',
        file_name=f'reporte_tarificacion_{id_tarificacion}.pdf',
        disposition='attachment'
    )
    message.add_attachment(pdf_attachment)
    
    # Adjuntar el XLS
    xls_attachment = Attachment(
        file_content=encoded_xls,
        file_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        file_name=f'reporte_tarificacion_{id_tarificacion}.xlsx',
        disposition='attachment'
    )
    message.add_attachment(xls_attachment)
    
    try:
        sg = SendGridAPIClient('')
        response = sg.send(message)
        return JsonResponse({'status': 'success', 'message': 'El reporte ha sido enviado por correo electrónico.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error al enviar el correo: {e}'})
      
@login_required
def generar_reportes(request):
    responsables = Usuario.objects.filter(rol_usuario='Responsable de Unidad')
    return render(request, 'generar_reportes.html', {'usuarios': responsables})

@login_required
def reporte_general(request):
    total_llamadas = Llamada.objects.count()
    total_segundos = Llamada.objects.aggregate(Sum('duracion_llamada'))['duracion_llamada__sum'] or 0
    total_minutos = total_segundos / 60  # Convertir a minutos
    promedio_duracion_segundos = Llamada.objects.aggregate(Avg('duracion_llamada'))['duracion_llamada__avg'] or 0
    promedio_duracion_minutos = promedio_duracion_segundos / 60  # Convertir a minutos

    llamadas = Llamada.objects.all()

    # Calcular el costo total de las llamadas
    costo_total = Decimal(0)
    for llamada in llamadas:
        proveedor = llamada.proveedor
        if llamada.tipo_llamada == 'CEL':
            tarifa = proveedor.tarifa_cel
        elif llamada.tipo_llamada == 'SLM':
            tarifa = proveedor.tarifa_slm
        elif llamada.tipo_llamada == 'LDI':
            tarifa = proveedor.tarifa_ldi
        else:
            tarifa = Decimal(0)  # Manejar el caso donde el tipo de llamada no coincide con ninguna tarifa

        costo_total += (Decimal(llamada.duracion_llamada) / Decimal(60)) * tarifa

    context = {
        'total_llamadas': total_llamadas,
        'total_minutos': total_minutos,
        'promedio_duracion': promedio_duracion_minutos,
        'costo_total': costo_total,
        'llamadas': llamadas,
    }

    if request.GET.get('download') == 'pdf':
        # Renderizar la plantilla HTML con los datos
        html_string = render_to_string('reporte_general.html', context)
        
        # Generar el PDF
        html = HTML(string=html_string)
        pdf = html.write_pdf()

        # Crear la respuesta HTTP con el PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_general.pdf"'
        
        return response

    return render(request, 'reporte_general.html', context)
  
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Sum, Avg
from decimal import Decimal
from .models import Llamada

def reporte_general_xls(request):
    try:
        total_llamadas = Llamada.objects.count()
        total_segundos = Llamada.objects.aggregate(Sum('duracion_llamada'))['duracion_llamada__sum'] or 0
        total_minutos = total_segundos / 60  # Convertir a minutos
        promedio_duracion_segundos = Llamada.objects.aggregate(Avg('duracion_llamada'))['duracion_llamada__avg'] or 0
        promedio_duracion_minutos = promedio_duracion_segundos / 60  # Convertir a minutos

        llamadas = Llamada.objects.all()

        # Calcular el costo total de las llamadas
        costo_total = Decimal(0)
        for llamada in llamadas:
            proveedor = llamada.proveedor
            if llamada.tipo_llamada == 'CEL':
                tarifa = proveedor.tarifa_cel
            elif llamada.tipo_llamada == 'SLM':
                tarifa = proveedor.tarifa_slm
            elif llamada.tipo_llamada == 'LDI':
                tarifa = proveedor.tarifa_ldi
            else:
                tarifa = Decimal(0)  # Manejar el caso donde el tipo de llamada no coincide con ninguna tarifa

            costo_total += (Decimal(llamada.duracion_llamada) / Decimal(60)) * tarifa

        # Generar el XLS
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte General de Llamadas"
        
        # Primera tabla: Reporte General de Llamadas
        ws.append(['', ''])
        title = "Reporte General de Llamadas"
        ws.merge_cells('B2:C2')
        cell = ws['B2']
        cell.value = title
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        headers = ['', 'Descripción', 'Valor']
        ws.append(headers)
        
        data = [
            ['', 'Total de Llamadas', total_llamadas],
            ['', 'Total de Minutos', total_minutos],
            ['', 'Promedio de Duración de Llamadas', promedio_duracion_minutos],
            ['', 'Costo Total de Llamadas', costo_total]
        ]
        
        for row in data:
            ws.append(row)
        
        # Añadir una fila vacía para la separación
        ws.append([''])
        
        # Segunda tabla: Detalle de Llamadas
        ws.append([''])  # Fila divisora en blanco
        
        detail_title = "Detalle de Llamadas"
        ws.merge_cells('B10:J10')
        cell = ws['B10']
        cell.value = detail_title
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        detail_headers = ['', 'Origen Llamada', 'Destino Llamada', 'Identificador Llamada', 'App Llamada', 'Inicio Llamada', 'Duración Llamada (segundos)', 'Segundos Facturados', 'Disposición', 'Tipo Llamada']
        ws.append(detail_headers)
        
        for llamada in llamadas:
            ws.append([
                '',
                llamada.origen_llamada,
                llamada.destino_llamada,
                llamada.identificador_llamada,
                llamada.app_llamada,
                llamada.inicio_llamada.strftime('%d/%m/%Y'),
                llamada.duracion_llamada,
                llamada.segundos_facturados,
                llamada.disposicion,
                llamada.tipo_llamada
            ])
        
        # Ajustar el ancho de las columnas y alinear el contenido para la primera tabla
        for column in ws.iter_cols(min_col=2, max_col=3, min_row=2, max_row=7):
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
                cell.alignment = Alignment(horizontal='center')
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Ajustar el ancho de las columnas y alinear el contenido para la segunda tabla
        for column in ws.iter_cols(min_col=2, max_col=10, min_row=10, max_row=ws.max_row):
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
                cell.alignment = Alignment(horizontal='center')
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        # Aplicar bordes a la primera tabla
        for row in ws.iter_rows(min_row=2, max_row=7, min_col=2, max_col=3):
            for cell in row:
                cell.border = thin_border
        
        # Aplicar bordes a la segunda tabla
        for row in ws.iter_rows(min_row=10, max_row=ws.max_row, min_col=2, max_col=10):
            for cell in row:
                cell.border = thin_border
        
        # Ajustar automáticamente el ancho de la columna B en función del contenido para evitar desbordamientos
        column_b_max_length = 0
        for row in range(2, ws.max_row + 1):
            cell_value_length = len(str(ws.cell(row=row, column=2).value))
            if cell_value_length > column_b_max_length:
                column_b_max_length = cell_value_length
        
        ws.column_dimensions['B'].width = column_b_max_length + 2
        
        xls_io = BytesIO()
        wb.save(xls_io)
        xls_io.seek(0)
        xls_data = xls_io.read()
        
        # Crear la respuesta HTTP con el archivo XLS
        response = HttpResponse(xls_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_general.xlsx"'
        
        return response

    except Exception as e:
        return HttpResponse(f'Error al generar el reporte: {e}', status=500)


@login_required
def enviar_reporte_general(request):
    total_llamadas = Llamada.objects.count()
    total_segundos = Llamada.objects.aggregate(Sum('duracion_llamada'))['duracion_llamada__sum'] or 0
    total_minutos = total_segundos / 60  # Convertir a minutos
    promedio_duracion_segundos = Llamada.objects.aggregate(Avg('duracion_llamada'))['duracion_llamada__avg'] or 0
    promedio_duracion_minutos = promedio_duracion_segundos / 60  # Convertir a minutos

    llamadas = Llamada.objects.all()

    # Calcular el costo total de las llamadas
    costo_total = Decimal(0)
    for llamada in llamadas:
        proveedor = llamada.proveedor
        if llamada.tipo_llamada == 'CEL':
            tarifa = proveedor.tarifa_cel
        elif llamada.tipo_llamada == 'SLM':
            tarifa = proveedor.tarifa_slm
        elif llamada.tipo_llamada == 'LDI':
            tarifa = proveedor.tarifa_ldi
        else:
            tarifa = Decimal(0)  # Manejar el caso donde el tipo de llamada no coincide con ninguna tarifa

        costo_total += (Decimal(llamada.duracion_llamada) / Decimal(60)) * tarifa

    context = {
        'total_llamadas': total_llamadas,
        'total_minutos': total_minutos,
        'promedio_duracion': promedio_duracion_minutos,
        'costo_total': costo_total,
        'llamadas': llamadas,
    }

    # Renderizar la plantilla HTML con los datos
    html_string = render_to_string('reporte_general.html', context)
    html = HTML(string=html_string)
    pdf = html.write_pdf()

    # Generar el XLS
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte General de Llamadas"
    
    # Primera tabla: Reporte General de Llamadas
    ws.append(['', ''])
    title = "Reporte General de Llamadas"
    ws.merge_cells('B2:C2')
    cell = ws['B2']
    cell.value = title
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    headers = ['', 'Descripción', 'Valor']
    ws.append(headers)
    
    data = [
        ['', 'Total de Llamadas', total_llamadas],
        ['', 'Total de Minutos', total_minutos],
        ['', 'Promedio de Duración de Llamadas', promedio_duracion_minutos],
        ['', 'Costo Total de Llamadas', costo_total]
    ]
    
    for row in data:
        ws.append(row)
    
    # Añadir una fila vacía para la separación
    ws.append([''])
    
    # Segunda tabla: Detalle de Llamadas
    ws.append([''])  # Fila divisora en blanco
    
    detail_title = "Detalle de Llamadas"
    ws.merge_cells('B10:J10')
    cell = ws['B10']
    cell.value = detail_title
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    detail_headers = ['', 'Origen Llamada', 'Destino Llamada', 'Identificador Llamada', 'App Llamada', 'Inicio Llamada', 'Duración Llamada (segundos)', 'Segundos Facturados', 'Disposición', 'Tipo Llamada']
    ws.append(detail_headers)
    
    for llamada in llamadas:
        ws.append([
            '',
            llamada.origen_llamada,
            llamada.destino_llamada,
            llamada.identificador_llamada,
            llamada.app_llamada,
            llamada.inicio_llamada.strftime('%d/%m/%Y'),
            llamada.duracion_llamada,
            llamada.segundos_facturados,
            llamada.disposicion,
            llamada.tipo_llamada
        ])
    
    # Ajustar el ancho de las columnas y alinear el contenido para la primera tabla
    for column in ws.iter_cols(min_col=2, max_col=3, min_row=2, max_row=7):
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            cell.alignment = Alignment(horizontal='center')
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Ajustar el ancho de las columnas y alinear el contenido para la segunda tabla
    for column in ws.iter_cols(min_col=2, max_col=10, min_row=10, max_row=ws.max_row):
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
              pass
            cell.alignment = Alignment(horizontal='center')
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Aplicar bordes a la primera tabla
    for row in ws.iter_rows(min_row=2, max_row=7, min_col=2, max_col=3):
        for cell in row:
            cell.border = thin_border
    
    # Aplicar bordes a la segunda tabla
    for row in ws.iter_rows(min_row=10, max_row=ws.max_row, min_col=2, max_col=10):
        for cell in row:
            cell.border = thin_border
    
    # Ajustar automáticamente el ancho de la columna B en función del contenido para evitar desbordamientos
    column_b_max_length = 0
    for row in range(2, ws.max_row + 1):
        cell_value_length = len(str(ws.cell(row=row, column=2).value))
        if cell_value_length > column_b_max_length:
            column_b_max_length = cell_value_length
    
    ws.column_dimensions['B'].width = column_b_max_length + 2
    
    xls_io = BytesIO()
    wb.save(xls_io)
    xls_io.seek(0)
    xls_data = xls_io.read()

    # Codificar los archivos en base64
    encoded_pdf = base64.b64encode(pdf).decode()
    encoded_xls = base64.b64encode(xls_data).decode()

    # Crear el mensaje de correo
    message = Mail(
        from_email='cr.barrera@duocuc.cl',
        to_emails=['cr.barrera@duocuc.cl', 'cristobal.24bn@gmail.com', 'cristobaljr24@gmail.com'],
        subject='Reporte General de Llamadas',
        html_content='Adjunto encontrarás el reporte general de llamadas solicitado.'
    )

    # Adjuntar el PDF
    pdf_attachment = Attachment(
        file_content=encoded_pdf,
        file_type='application/pdf',
        file_name='reporte_general.pdf',
        disposition='attachment'
    )
    message.add_attachment(pdf_attachment)

    # Adjuntar el XLS
    xls_attachment = Attachment(
        file_content=encoded_xls,
        file_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        file_name='reporte_general.xlsx',
        disposition='attachment'
    )
    message.add_attachment(xls_attachment)

    try:
        sg = SendGridAPIClient('')
        response = sg.send(message)
        return JsonResponse({'status': 'success', 'message': 'El reporte ha sido enviado por correo electrónico.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error al enviar el correo: {e}'})

@login_required 
def generar_reporte_responsable_pdf(request, username):
    usuario = get_object_or_404(Usuario, username=username)
    anexos = Anexo.objects.filter(usuario=usuario)
    tarificaciones = Tarificacion.objects.filter(anexo__in=anexos)

    # Renderizar la plantilla HTML con los datos de las tarificaciones
    html_string = render_to_string('reporte_responsable_unidad.html', {'usuario': usuario, 'tarificaciones': tarificaciones})
    
    # Generar el PDF
    html = HTML(string=html_string)
    pdf = html.write_pdf()

    # Crear la respuesta HTTP con el PDF
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_responsable_{username}.pdf"'
    
    return response
  
def generar_reporte_responsable_xls(request, username):
    usuario = get_object_or_404(Usuario, username=username)
    anexos = Anexo.objects.filter(usuario=usuario)
    tarificaciones = Tarificacion.objects.filter(anexo__in=anexos)

    # Crear un libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Tarificación"

    # Insertar una fila vacía al inicio
    ws.append([''])

    # Añadir el título y fusionar celdas
    title = f"Reporte de Tarificación para {usuario.username}"
    ws.merge_cells('B2:G2')
    cell = ws['B2']
    cell.value = title
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Añadir bordes al título
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=7):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

    # Añadir los encabezados con una columna vacía al inicio
    headers = ['', 'ID Tarificación', 'Fecha Inicio', 'Fecha Término', 'Costo Total', 'Minutos Totales', 'Tipo Llamada']
    ws.append(headers)

    # Añadir los datos de las tarificaciones con una columna vacía al inicio
    for tarificacion in tarificaciones:
        row = [
            '',
            tarificacion.id_tarificacion,
            tarificacion.fecha_inicio.strftime('%d/%m/%Y'),
            tarificacion.fecha_termino.strftime('%d/%m/%Y'),
            tarificacion.costo_total,
            tarificacion.minutos_totales,
            tarificacion.tipo_llamada
        ]
        ws.append(row)

    # Ajustar el ancho de las columnas y alinear el contenido
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[column_letter].width = max_length + 2

    # Añadir bordes a las celdas
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=7):
        for cell in row:
            cell.border = thin_border

    # Crear la respuesta HTTP con el archivo XLS
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_responsable_{username}.xlsx"'

    # Guardar el libro de trabajo en la respuesta
    wb.save(response)

    return response
  
@login_required
def consultar_trafico_por_proveedor(request):
    trafico_por_proveedor = Llamada.objects.values('proveedor__nombre_proveedor').annotate(
        total_duracion=Sum('duracion_llamada'),
        total_minutos_facturados=Sum(ExpressionWrapper(F('segundos_facturados') / 60, output_field=fields.FloatField()))
    ).order_by('proveedor__nombre_proveedor')

    context = {
        'trafico_por_proveedor': trafico_por_proveedor
    }
    return render(request, 'consultar_trafico_por_proveedor.html', context)